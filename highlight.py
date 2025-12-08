import os.path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook
#file_path = r"C:\SAP Project\Reports\Aug 2025\PC_summary August 01_04.xlsx"  # Replace with your file path
def hihglight(file_path,file):
    wb = load_workbook(os.path.join(file_path,file))

    # Define red fill for highlighting
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Loop through all sheets and cells
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            # Check if any cell in the row contains "failed" (case-insensitive)
            if any(cell.value and isinstance(cell.value, str) and "failed" in cell.value.lower() for cell in row):
                # Highlight the full row
                for cell in row:
                    cell.fill = red_fill

    # Save the workbook
    wb.save(os.path.join(file_path,file))
    print("Cells with 'failed' have been highlighted in red.")