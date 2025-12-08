import os
import cv2
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
def image_to_text(input_filename,output_filename):
    try:

        # === Load Excel workbook and worksheet ===

        #input_filename="Process_chain_23_june_2025.xlsx"
        wb = load_workbook(input_filename)
        ws = wb.active

        # Create temporary folder to save images
        os.makedirs("extracted_images", exist_ok=True)

        # Get list of all images
        images = ws._images[:]

        def detect_color(image):
            """Detect red, yellow, or default to success."""
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

            # Red color mask (two ranges)
            red1 = cv2.inRange(hsv, (0, 100, 100), (10, 255, 255))
            red2 = cv2.inRange(hsv, (160, 100, 100), (180, 255, 255))
            red_mask = cv2.bitwise_or(red1, red2)

            # Improved yellow mask (broader)
            yellow_mask = cv2.inRange(hsv, (15, 80, 80), (40, 255, 255))

            red_pixels = cv2.countNonZero(red_mask)
            yellow_pixels = cv2.countNonZero(yellow_mask)
            total_pixels = image.shape[0] * image.shape[1]

            red_ratio = red_pixels / total_pixels
            yellow_ratio = yellow_pixels / total_pixels

            print(f"[DEBUG] Red: {red_ratio:.2%}, Yellow: {yellow_ratio:.2%}")

            if red_ratio > 0.04:
                return "Failed"
            elif yellow_ratio > 0.04:
                return "Running"
            else:
                return "Success"

        # === Process each image ===
        for idx, img in enumerate(images):
            anchor = img.anchor._from
            col_letter = get_column_letter(anchor.col + 1)
            row_number = anchor.row + 1
            cell_address = f"{col_letter}{row_number}"

            # ✅ Skip if cell is blank (before replacing)
            if ws[cell_address].value is None:
                print(f"Skipping blank cell at {cell_address}")
                continue

            # Save image temporarily
            image_path = f"extracted_images/img_{idx+1}_{cell_address}.png"
            with open(image_path, "wb") as f:
                f.write(img._data())

            # Load image using OpenCV
            image = cv2.imread(image_path)
            if image is None:
                print(f"Could not read image at {cell_address}")
                continue

            # Detect color
            result_text = detect_color(image)

            # Replace image with result
            ws[cell_address] = result_text
            ws._images.remove(img)
            #ws[cell_address].alignment = Alignment(horizontal='left')
            print(f"Replaced image at {cell_address} with '{result_text}'")

        # === Save the updated workbook ===
        #output_directory=r"C:\SAP Project\Reports"
        #output_filename="Process_chain_23_june_2025_final.xlsx"
        wb.save(output_filename)
        print("✅ Done. Saved as "+output_filename)
        #return output_filename
    except Exception as e:
        print(e)
